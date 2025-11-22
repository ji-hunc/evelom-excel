import streamlit as st
import io
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from collections import defaultdict

st.title("Excel 처리 앱")
st.write("""
1️⃣ 엑셀 파일을 업로드하면,  
2️⃣ 가공된 파일을 다운로드할 수 있어요!
""")

uploaded_file = st.file_uploader("엑셀 파일 업로드 (.xlsx)", type=["xlsx"], accept_multiple_files=False)

if uploaded_file:
    if st.button("✅ 처리 시작하기"):
        try:
            # 파일명에서 날짜 추출하여 결과 파일명 설정
            uploaded_filename = uploaded_file.name
            match = re.search(r"\d{8}", uploaded_filename)
            if match:
                extracted_date = match.group()
                output_filename = f"결과_{extracted_date}.xlsx"
            else:
                output_filename = "processed_excel.xlsx"

            progress_bar = st.progress(0)
            st.info("파일 처리 중... 잠시만 기다려주세요.")

            # 1. openpyxl로 파일 열기
            wb = load_workbook(uploaded_file)
            ws = wb.active
            progress_bar.progress(10)

            # 2. 전체 셀 배경색 제거 (1행은 제외)
            no_fill = PatternFill(fill_type=None)
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                if row_idx == 1:
                    continue
                for cell in row:
                    cell.fill = no_fill
            progress_bar.progress(20)

            # 3. 열 너비 조정
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['C'].width = 80
            ws.column_dimensions['E'].width = 5
            ws.column_dimensions['G'].width = 30
            ws.column_dimensions['H'].width = 30
            ws.column_dimensions['I'].width = 7
            ws.column_dimensions['J'].width = 25
            ws.column_dimensions['F'].width = 25
            ws.column_dimensions['L'].width = 80
            progress_bar.progress(30)

            # ===== 4. 수취인 기준 그룹 만들기 =====
            max_col = ws.max_column
            original_max_row = ws.max_row

            groups = []
            current_group = []
            current_recipient = None

            for row_idx in range(2, original_max_row + 1):
                row_values = [
                    ws.cell(row=row_idx, column=col_idx).value
                    for col_idx in range(1, max_col + 1)
                ]
                recipient = row_values[1]  # B열 (수취인)

                if recipient is None or str(recipient).strip() == "":
                    if current_group:
                        groups.append(current_group)
                        current_group = []
                        current_recipient = None
                    continue

                if current_recipient is None:
                    current_recipient = recipient
                elif recipient != current_recipient:
                    groups.append(current_group)
                    current_group = []
                    current_recipient = recipient

                current_group.append(row_values)

            if current_group:
                groups.append(current_group)

            # ===== 그룹 내부 상품명(C열) 정렬 =====
            group_objs = []

            for g in groups:
                sorted_rows = sorted(
                    g,
                    key=lambda row: str(row[2]) if row[2] is not None else ""
                )

                if sorted_rows and sorted_rows[0][2] is not None:
                    group_key = str(sorted_rows[0][2])
                else:
                    group_key = ""

                group_objs.append({
                    "rows": sorted_rows,
                    "key": group_key
                })

            # ===== 그룹 자체도 상품명 기준으로 정렬 =====
            group_objs.sort(key=lambda g: g["key"])

            # 데이터 영역 초기화
            for row_idx in range(2, original_max_row + 1):
                for col_idx in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).value = None
                    ws.cell(row=row_idx, column=col_idx).fill = no_fill

            # ===== 정렬된 그룹 다시 쓰기 + 회색 구분선 =====
            light_gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            max_col_to_fill = max(max_col, 50)

            write_row = 2

            for gi, g in enumerate(group_objs):
                if gi > 0:
                    for col_idx in range(1, max_col_to_fill + 1):
                        ws.cell(row=write_row, column=col_idx).fill = light_gray_fill
                    write_row += 1

                for row_values in g["rows"]:
                    for col_idx, value in enumerate(row_values, start=1):
                        ws.cell(row=write_row, column=col_idx).value = value
                    write_row += 1

            progress_bar.progress(60)

            # ===== 4-1. 전체 본문 border 설정 =====
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for row_idx in range(2, ws.max_row + 1):
                for col_idx in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).border = thin_border

            # ===== 4-2. Row height 구분 적용 =====
            for row_idx in range(2, ws.max_row + 1):

                # 구분선 판별 (회색 fill)
                first_cell = ws.cell(row=row_idx, column=1)
                fill = first_cell.fill

                is_separator = (
                    fill.patternType == "solid" and 
                    (fill.start_color.rgb == "00E0E0E0" or fill.start_color.rgb == "FFE0E0E0")
                )

                if is_separator:
                    ws.row_dimensions[row_idx].height = 15  # 구분선 row
                else:
                    ws.row_dimensions[row_idx].height = 30  # 일반 데이터 row

            # ===== 4-3. 모든 셀 가운데 정렬 =====
            center_align = Alignment(vertical="center")

            for row_idx in range(1, ws.max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = center_align



            # ===== 5. 핑크 색칠 =====
            pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")

            for row_idx in range(2, ws.max_row + 1):
                for col_idx in [5, 9]:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    if value is None:
                        continue

                    cell.alignment = center_align

                    try:
                        numeric_value = float(str(value).strip())
                        if numeric_value >= 2:
                            cell.fill = pink_fill
                    except:
                        pass

            progress_bar.progress(80)

            # ===== 6. 상품명별 합계 =====
            product_sales = defaultdict(float)

            for row_idx in range(2, ws.max_row + 1):
                product = ws.cell(row=row_idx, column=3).value
                quantity = ws.cell(row=row_idx, column=5).value

                if product is None:
                    continue

                try:
                    quantity_num = float(str(quantity).strip()) if quantity else 0
                except:
                    quantity_num = 0

                product_sales[str(product).strip()] += quantity_num

            summary_start_row = ws.max_row + 6
            gothic_font = Font(name='맑은 고딕', size=10)

            for i, (product, total_qty) in enumerate(product_sales.items()):
                row_idx = summary_start_row + i

                ws.cell(row=row_idx, column=3).value = product
                ws.cell(row=row_idx, column=3).border = thin_border
                ws.cell(row=row_idx, column=3).font = gothic_font

                ws.cell(row=row_idx, column=5).value = total_qty
                ws.cell(row=row_idx, column=5).border = thin_border
                ws.cell(row=row_idx, column=5).font = gothic_font

                ws.cell(row=row_idx, column=6).alignment = center_align

                ws.row_dimensions[row_idx].height = 25

            progress_bar.progress(95)

            # 7. 파일 저장
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            progress_bar.progress(100)

            st.success("🎉 처리 완료! 아래 버튼으로 다운로드하세요.")
            st.download_button(
                label="📥 가공된 엑셀 다운로드",
                data=output,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        except Exception as e:
            st.error(f"⚠️ 처리 중 오류 발생: {e}")
